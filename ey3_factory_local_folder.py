#!/usr/bin/env python
# coding=UTF-8
#
from openpyxl import Workbook
import glob

FOLDER_PREFIX_NAME = "BEY"

FILENAME = "ey3_factory.xlsx"

ROW_OFFSET = 1
MTF_START_POS1 = 2
MTF_1_LEN = 9
MTF_START_POS2 = (MTF_START_POS1 + MTF_1_LEN)
MTF_2_LEN = 9
SFR_START_POS = (MTF_START_POS2 + MTF_2_LEN)
SFR_LEN = 36
TP1_START_POS1 = (SFR_START_POS + SFR_LEN)
TP1_1_LEN = 3
TP1_START_POS2 = (TP1_START_POS1 + TP1_1_LEN)
TP1_2_LEN = 42
TP2_START_POS = (TP1_START_POS2 + TP1_2_LEN)
TP2_LEN = 13
COLOR_FIDELITY_START_POS = (TP2_START_POS + TP2_LEN)
COLOR_FIDELITY_LEN = 12


def excel_create():
    wb = Workbook()
    return wb


def excel_save(wb, filename):
    # Save the file
    wb.save(filename)


def excel_creatsheet(wb, ws_title):
    ws = wb.create_sheet(title=ws_title)
    return ws


def find_directoies_with_substring(ey):
    return glob.glob(ey)


def mtf_dealwith(ws, ey_folder, ey_folder_idx):
    ws.cell(column=1, row=ey_folder_idx + ROW_OFFSET, value=ey_folder)
    filefullpath = ey_folder + "/mtf/" + ey_folder + "-H-MTFOUT.txt"
    try:
        with open(filefullpath, "r") as f:
            for mtf_idx, line in enumerate(f, MTF_START_POS1):
                roi, mtf = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=mtf_idx, row=1, value=roi)
                ws.cell(column=mtf_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(mtf))
    except FileNotFoundError:
        pass
    filefullpath = ey_folder + "/mtf/" + ey_folder + "-V-MTFOUT.txt"
    try:
        with open(filefullpath, "r") as f:
            for mtf_idx, line in enumerate(f, MTF_START_POS2):
                roi, mtf = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=mtf_idx, row=1, value=roi)
                ws.cell(column=mtf_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(mtf))
    except FileNotFoundError:
        pass


def sfr_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = ey_folder + "/sfr/" + "SFROUT_shopfloor.txt"
    try:
        with open(filefullpath, "r") as f:
            for sfr_idx, line in enumerate(f, SFR_START_POS):
                roi, sfr = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=sfr_idx, row=1, value=roi)
                ws.cell(column=sfr_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(sfr))
    except FileNotFoundError:
        pass


def tp1_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = ey_folder + "/tp1/" + "tp1-black-out.txt"
    try:
        with open(filefullpath, "r") as f:
            for tp1_idx, line in enumerate(f, TP1_START_POS1):
                roi, black = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=tp1_idx, row=1, value=roi)
                ws.cell(column=tp1_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(black))
    except FileNotFoundError:
        pass
    try:
        filefullpath = ey_folder + "/tp1/" + "tp1-white-out.txt"
        with open(filefullpath, "r") as f:
            for tp1_idx, line in enumerate(f, TP1_START_POS2):
                roi, white = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=tp1_idx, row=1, value=roi)
                ws.cell(column=tp1_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(white))
    except FileNotFoundError:
        pass


def tp2_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = ey_folder + "/tp2/" + "chart-out.txt"
    try:
        with open(filefullpath, "r") as f:
            for tp2_idx, line in enumerate(f, TP2_START_POS):
                roi, tp2 = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=tp2_idx, row=1, value=roi)
                ws.cell(column=tp2_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(tp2))
    except FileNotFoundError:
        pass


def color_fidelity_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = ey_folder + "/" + "color_fidelity.txt"
    try:
        with open(filefullpath, "r") as f:
            for cfd_idx, line in enumerate(f, COLOR_FIDELITY_START_POS):
                roi, cfd = line.split('=')
                if ey_folder_idx == 1:
                    ws.cell(column=cfd_idx, row=1, value=roi)
                ws.cell(column=cfd_idx, row=ey_folder_idx + ROW_OFFSET,
                        value=float(cfd))
    except FileNotFoundError:
        pass


if __name__ == '__main__':
    wb = excel_create()
    ws = wb.active
    EY_FOLDERS = find_directoies_with_substring(FOLDER_PREFIX_NAME + "*")
    for ey_folder_idx, ey_folder in enumerate(EY_FOLDERS, 1):
        mtf_dealwith(ws, ey_folder, ey_folder_idx)
        sfr_dealwith(ws, ey_folder, ey_folder_idx)
        tp1_dealwith(ws, ey_folder, ey_folder_idx)
        tp2_dealwith(ws, ey_folder, ey_folder_idx)
        color_fidelity_dealwith(ws, ey_folder, ey_folder_idx)
    excel_save(wb, FILENAME)
