#!/usr/bin/env python
# coding=UTF-8
from openpyxl import Workbook
import glob
#
# Copyright (C) 2016 Quanta Computer Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# increment this whenever we make important changes to this script
VERSION = (0, 2)


# Maintainer <jim_lin@quantatw.com>


FILENAME = "ey3_factory.xlsx"
ROW_OFFSET = 1
CHART_OUT_POS = 2
CHART_OUT_LEN = 13
COLOR_FIDELITY_START_POS = (CHART_OUT_POS + CHART_OUT_LEN)
COLOR_FIDELITY_LEN = 12
MTF_START_POS1 = (COLOR_FIDELITY_START_POS + COLOR_FIDELITY_LEN)
MTF_1_LEN = 9
MTF_START_POS2 = (MTF_START_POS1 + MTF_1_LEN)
MTF_2_LEN = 9
SFR_START_POS = (MTF_START_POS2 + MTF_2_LEN)
SFR_LEN = 36
BLACK_START_POS = (SFR_START_POS + SFR_LEN)
BLACK_LEN = 3
DIFFUSER_1_5000_POS = (BLACK_START_POS + BLACK_LEN)
DIFFUSER_1_5000_LEN = 42
DIFFUSER_2_3000_POS = (DIFFUSER_1_5000_POS + DIFFUSER_1_5000_LEN)
DIFFUSER_2_3000_LEN = 42
CAMERA_STATION_1 = "Camera1"
CAMERA_STATION_2 = "Camera2"
MATCH_PATTEN = CAMERA_STATION_1 + "/*-black-0-/*"


def excel_create():
    wb = Workbook()
    return wb


def excel_save(wb, filename):
    # Save the file
    wb.save(filename)


def excel_creatsheet(wb, ws_title):
    ws = wb.create_sheet(title=ws_title)
    return ws


def find_directoies_with_substring(ey):
    return glob.glob(ey)


def mtf_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = CAMERA_STATION_2 + "/" + ey_folder + \
        "-chart-0-/mtf-H-MTFOUT.txt"
    with open(filefullpath, "r") as f:
        for mtf_idx, line in enumerate(f, MTF_START_POS1):
            roi, mtf = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=mtf_idx, row=1, value=roi)
            ws.cell(column=mtf_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(mtf))
    filefullpath = CAMERA_STATION_2 + "/" + ey_folder + \
        "-chart-0-/mtf-V-MTFOUT.txt"
    with open(filefullpath, "r") as f:
        for mtf_idx, line in enumerate(f, MTF_START_POS2):
            roi, mtf = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=mtf_idx, row=1, value=roi)
            ws.cell(column=mtf_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(mtf))


def sfr_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = CAMERA_STATION_2 + "/" + ey_folder + \
        "-chart-0-/SFROUT_shopfloor.txt"
    with open(filefullpath, "r") as f:
        for sfr_idx, line in enumerate(f, SFR_START_POS):
            roi, sfr = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=sfr_idx, row=1, value=roi)
            ws.cell(column=sfr_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(sfr))


def chart_out_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = CAMERA_STATION_2 + "/" + ey_folder + \
        "-chart-0-/chart-out.txt"
    with open(filefullpath, "r") as f:
        for co_idx, line in enumerate(f, CHART_OUT_POS):
            roi, co = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=co_idx, row=1, value=roi)
            ws.cell(column=co_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(co))


def color_fidelity_dealwith(ws, ey_folder, ey_folder_idx):
    filefullpath = CAMERA_STATION_2 + "/" + ey_folder + \
        "-chart-0-/color_fidelity.txt"
    with open(filefullpath, "r") as f:
        for cfd_idx, line in enumerate(f, COLOR_FIDELITY_START_POS):
            roi, cfd = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=cfd_idx, row=1, value=roi)
            ws.cell(column=cfd_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(cfd))


def black_0_dealwith(ws, ey_folder, ey_folder_idx):

    ws.cell(column=1, row=ey_folder_idx + ROW_OFFSET, value=ey_folder)
    filefullpath = CAMERA_STATION_1 + "/" + ey_folder + \
        "-black-0-/tp1-black-out.txt"
    with open(filefullpath, "r") as f:
        for black_idx, line in enumerate(f, BLACK_START_POS):
            roi, black = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=black_idx, row=1, value=roi)
            ws.cell(column=black_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(black))


def diffuser_1_5000_dealwith(ws, ey_folder, ey_folder_idx):

    filefullpath = CAMERA_STATION_1 + "/" + ey_folder + \
        "-diffuser-1-5000/tp1-white-out.txt"
    with open(filefullpath, "r") as f:
        for du1_idx, line in enumerate(f, DIFFUSER_1_5000_POS):
            roi, du1 = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=du1_idx, row=1, value=roi)
            ws.cell(column=du1_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(du1))


def diffuser_2_3000_dealwith(ws, ey_folder, ey_folder_idx):

    filefullpath = CAMERA_STATION_1 + "/" + ey_folder + \
        "-diffuser-2-3000/tp1-white-out.txt"
    with open(filefullpath, "r") as f:
        for du2_idx, line in enumerate(f, DIFFUSER_2_3000_POS):
            roi, du2 = line.split('=')
            if ey_folder_idx == 1:
                ws.cell(column=du2_idx, row=1, value=roi)
            ws.cell(column=du2_idx, row=ey_folder_idx + ROW_OFFSET,
                    value=float(du2))


if __name__ == '__main__':
    EY_FOLDERS = []
    wb = excel_create()
    ws = wb.active
    for ey_folder in find_directoies_with_substring(MATCH_PATTEN):
        EY_FOLDERS.append(ey_folder[8:20])
        # print(EY_FOLDERS)
    for ey_folder_idx, ey_folder in enumerate(EY_FOLDERS, 1):
        chart_out_dealwith(ws, ey_folder, ey_folder_idx)
        color_fidelity_dealwith(ws, ey_folder, ey_folder_idx)
        mtf_dealwith(ws, ey_folder, ey_folder_idx)
        sfr_dealwith(ws, ey_folder, ey_folder_idx)
        black_0_dealwith(ws, ey_folder, ey_folder_idx)
        diffuser_1_5000_dealwith(ws, ey_folder, ey_folder_idx)
    excel_save(wb, FILENAME)
