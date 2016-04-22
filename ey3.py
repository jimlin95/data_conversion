#!/usr/bin/env python
# coding=UTF-8
from openpyxl import Workbook
from openpyxl.styles import (
        Side,
        Font,
        PatternFill,
        Border,
        Alignment,
        fills,
        colors
    )
from openpyxl.chart import Reference, LineChart
import glob

#
# Copyright (C) 2016 Quanta Computer Inc.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# increment this whenever we make important changes to this script
VERSION = (0, 2)


# Maintainer <jim_lin@quantatw.com>


FOLDER_PREFIX_NAME = "*EY"
FILENAME = "ey3.xlsx"
SHEET_SFR = "SFR"
SHEET_SFR_2 = "SFR_2"
SHEET_CTF = "CTF"
AREA_TAG = ("UL-0.5", "UR-0.5", "UL-0.3", "UR-0.3", "Center",
            "LL-0.3", "LR-0.3", "LL-0.5", "LR-0.5")
SFR2_TAG = ("0.5f", "0.3f", "0", "0.3f", "0.5f")
EY_FOLDERS = []
EY_FOLDERS_NUM = 0
CTF_INITIATE_TABLE = (
        ("cy/mm", "ROI"),
        (77.265, "ROI_0_MTF"),
        (57.433, "ROI_1_MTF"),
        (56.759, "ROI_2_MTF"),
        (74.585, "ROI_3_MTF"),
        (52.066, "ROI_9_MTF"),
        (56.428, "ROI_13_MTF"),
        (56.759, "ROI_14_MTF"),
        (75.679, "ROI_16_MTF"),
        (73.523, "ROI_17_MTF"),
        (71.486, "ROI_4_MTF"),
        (70.703, "ROI_5_MTF"),
        (57.263, "ROI_6_MTF"),
        (56.759, "ROI_7_MTF"),
        (51.523, "ROI_8_MTF"),
        (56.926, "ROI_10_MTF"),
        (56.593, "ROI_11_MTF"),
        (71.486, "ROI_12_MTF"),
        (70.317, "ROI_15_MTF"),
    )

TGT_FREQ = (
        (77.265),
        (57.433),
        (56.759),
        (74.585),
        (52.066),
        (56.428),
        (56.759),
        (75.679),
        (73.523),
        (71.486),
        (70.703),
        (57.263),
        (56.759),
        (51.523),
        (56.926),
        (56.593),
        (71.486),
        (70.317)
)

ROI_table = (
    "ROI_0_MTF",
    "ROI_1_MTF",
    "ROI_2_MTF",
    "ROI_3_MTF",
    "ROI_9_MTF",
    "ROI_13_MTF",
    "ROI_14_MTF",
    "ROI_16_MTF",
    "ROI_17_MTF",
    "ROI_4_MTF",
    "ROI_5_MTF",
    "ROI_6_MTF",
    "ROI_7_MTF",
    "ROI_8_MTF",
    "ROI_10_MTF",
    "ROI_11_MTF",
    "ROI_12_MTF",
    "ROI_15_MTF"
)


def find_between(s, first, last):
    try:
        start = s.index(first) + len(first)
        end = s.index(last, start)
        return s[start:end]
    except ValueError:
        return ""


def find_between_r(s, first, last):
    try:
        start = s.rindex(first) + len(first)
        end = s.rindex(last, start)
        return s[start:end]
    except ValueError:
        return ""


def excel_create():
    wb = Workbook()
    ws = wb.active
    # remove the sheet named "Sheet"
    wb.remove_sheet(ws)
    return wb


def excel_save(wb, filename):
    # Save the file
    wb.save(filename)


def set_allborder(ws, cell_range):
    rows = list(ws.iter_rows(cell_range))
    side = Side(border_style='thin', color="FF000000")
    for pos_y, cells in enumerate(rows):
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            border.left = side
            border.right = side
            border.top = side
            border.bottom = side
            cell.border = border


def set_border(ws, cell_range):
    rows = list(ws.iter_rows(cell_range))
    side = Side(border_style='thin', color="FF000000")
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def set_font(ws, cell_range):
    rows = list(ws.iter_rows(cell_range))
    font = Font(name='Calibri', size=11, bold=False, italic=False,
                vertAlign=None, underline='none', strike=False,
                color='FFFFFFFF')

    for pos_y, cells in enumerate(rows):
        for pos_x, cell in enumerate(cells):
            cell.font = font


def set_alignment(ws, cell_range, position='center'):
    rows = list(ws.iter_rows(cell_range))
    align_center = Alignment(horizontal=position)
    for pos_y, cells in enumerate(rows):
        for pos_x, cell in enumerate(cells):
            cell.alignment = align_center


def set_background_color(ws, cell_range, color_string):
    rows = list(ws.iter_rows(cell_range))
    backgroundcolor = PatternFill(
        fill_type="solid",
        start_color='FF' + color_string,
        end_color='FF' + color_string)
    for pos_y, cells in enumerate(rows):
        for pos_x, cell in enumerate(cells):
            cell.fill = backgroundcolor


def sfr_dealwith(ws, eyfile_index):
    adjust_index = 3 * (eyfile_index) + 2
    ws.cell(column=adjust_index, row=1, value=EY_FOLDERS[eyfile_index])
    ws.cell(column=adjust_index+2, row=1, value=EY_FOLDERS[eyfile_index])
    filefullpath = EY_FOLDERS[eyfile_index] + "/sfr/" + "SFROUT_shopfloor.txt"
    with open(filefullpath, "r") as f:
        for line in f:
            roi, sfr = line.split('=')
            index = find_between(roi, "ROI", "_SFR_RESULT")
            ws.cell(column=1, row=int(index)+2, value=roi)
            ws.cell(column=adjust_index, row=int(index)+2, value=float(sfr))
            for ul_index, i in enumerate(range(5, 38, 4)):
                ws.cell(column=3*(eyfile_index+1), row=i,
                        value=AREA_TAG[ul_index])
                # write SUM(4 values)/4 in UL/UR/LL/LR cell
                pos = chr(ord('B') + eyfile_index*3)
                ws.cell(column=(3*(eyfile_index+1))+1, row=i,
                        value="=SUM({0}{1}:{2}{3})/4".format(pos, i-3, pos, i))
                # paint in blue and green
                color_string = 'C6D9F1'  # blue color hex string
                datarow = chr(ord('C') + 3 * (eyfile_index))
                set_background_color(ws, "{0}2:{0}9".format(datarow),
                                     color_string)
                set_background_color(ws, "{0}30:{0}37".format(datarow),
                                     color_string)
                color_string = 'C3D69B'  # green color
                set_background_color(ws, "{0}10:{0}17".format(datarow),
                                     color_string)
                set_background_color(ws, "{0}22:{0}29".format(datarow),
                                     color_string)


def sfr_linechart(ws):
    chart1 = LineChart()
    chart1.title = "Line Chart"
    chart1.style = 2
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''

    for idx, i in enumerate(range(2, 3*EY_FOLDERS_NUM, 3)):
        # Select all data include title
        data = Reference(ws, min_col=i, min_row=1, max_row=37, max_col=i)
        chart1.add_data(data, titles_from_data=True)
        s1 = chart1.series[idx]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        # Marker outline
        s1.marker.graphicalProperties.line.solidFill = "FF0000"
        s1.graphicalProperties.line.noFill = False
    ws.add_chart(chart1, "H5")


def sfr_handle(ws):
    "Deal with all SFR files"
    global EY_FOLDERS
    global EY_FOLDERS_NUM
    set_alignment(ws, "A1:AZ1")
    EY_FOLDERS_TEMP = find_directoies_with_substring(FOLDER_PREFIX_NAME + "*/")
    for ey_folder in EY_FOLDERS_TEMP:
        EY_FOLDERS.append(ey_folder[:-1])
    # print(EY_FOLDERS)
    for index, folder in enumerate(EY_FOLDERS):
        sfr_dealwith(ws, index)
    EY_FOLDERS_NUM = len(EY_FOLDERS)
    sfr_linechart(ws)


def sfr2_linechart(ws):
    chart1 = LineChart()
    chart1.title = "Line Chart"
    chart1.style = 2
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''

    for idx in range(EY_FOLDERS_NUM):
        # Select all data include title
        data = Reference(ws, min_col=3+idx, min_row=1, max_row=10,
                         max_col=3+idx)
        chart1.add_data(data, titles_from_data=True)
        s1 = chart1.series[idx]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        # Marker outline
        s1.marker.graphicalProperties.line.solidFill = "FF0000"
        s1.graphicalProperties.line.noFill = False
    ws.add_chart(chart1, "B22")


def copy_sfr_to_sfr2(wb):
    "Copy UL datas to SFR2 sheet from SFR sheet"
    SFR_ws = wb[SHEET_SFR]
    SFR2_ws = wb[SHEET_SFR_2]
    set_alignment(SFR2_ws, "A1:AZ1")
    # Currently , we calculate average here, need to find how to read real value
    # by openpyxl
    for idx in range(EY_FOLDERS_NUM):  # 0, 1, 2, ..
        all_avg_value = []
        for y_idx in range(2, 38, 4):  # 2, 6, 10, 14, 18, 22, 26, 30, 34
            sum_4value = 0
            for sum_idx in range(4):  # 0, 1, 2, 3
                sum_pos = "{0}{1}".format(chr(ord('B') + idx*3), y_idx+sum_idx)
                sum_4value += float("{0:.7f}".format(SFR_ws[sum_pos].value))
            all_avg_value.append("{0:.7f}".format(sum_4value/4))
        sfr_title_pos = "{0}{1}".format(chr(ord('B') + idx*3), 1)
        sfr2_title_pos = "{0}{1}".format(chr(ord('C') + idx), 1)
        SFR2_ws[sfr2_title_pos] = SFR_ws[sfr_title_pos].value
        for avg_idx, sfr2_idx in enumerate(range(2, 11)):
            pos = "{0}{1}".format(chr(ord('C') + idx), sfr2_idx)
            SFR2_ws[pos] = float(all_avg_value[avg_idx])


def excel_creatsheet(wb, ws_title):
    ws = wb.create_sheet(title=ws_title)
    return ws


def create_working_sheets(wb):
    SHEETS_LIST = (SHEET_SFR, SHEET_SFR_2, SHEET_CTF)
    worksheets = []
    for sheet in SHEETS_LIST:
        worksheets.append(excel_creatsheet(wb, sheet))
    # print(worksheets)


def find_directoies_with_substring(ey):
    return glob.glob(ey)


def sfr_sheet_initiate(ws):
    "A function to initiate SFR sheet"
    # paste "ROI0_SFR_RESULT" to A2:A37
    for index, row_index in enumerate(range(2, 38)):
        ws.cell(column=1, row=row_index,
                value="ROI{0}_SFR_RESULT".format(index))
        ws.column_dimensions['A'].width = 18


def sfr2_sheet_initiate(ws):
    "A function to initiate SFR-2 sheet"
    for index, i in enumerate(range(2, 11)):
        # format column B2:B10
        ws.cell(column=2, row=i).style.fill.fill_type = fills.FILL_SOLID
        ws.cell(column=2, row=i).style.fill.start_color = colors.DARKRED
        ws.cell(column=2, row=i).value = AREA_TAG[index]
    for index, i in enumerate(range(13, 22)):
        ul = 'C' + str(i)
        ws[ul] = AREA_TAG[index]
    for f_index, f in enumerate(range(6, 11)):
        ws.cell(column=f, row=13, value=SFR2_TAG[f_index])
    for f_index, f in enumerate(range(14, 19)):
        ws.cell(column=5, row=f, value=SFR2_TAG[f_index])
    ws['F14'] = "=D13"
    ws['J14'] = "=D14"
    ws['G15'] = "=D15"
    ws['I15'] = "=D16"
    ws['H16'] = "=D17"
    ws['G17'] = "=D18"
    ws['I17'] = "=D19"
    ws['F18'] = "=D20"
    ws['J18'] = "=D21"
    set_alignment(ws, "E13:J18")
    set_allborder(ws, "E13:J18")
    color_string = 'C6D9F1'  # blue color hex string
    set_background_color(ws, "B2:B3", color_string)
    set_background_color(ws, "B9:B10", color_string)
    set_background_color(ws, "C13:C14", color_string)
    set_background_color(ws, "C20:C21", color_string)
    color_string = 'C3D69B'  # green color
    set_background_color(ws, "B4:B5", color_string)
    set_background_color(ws, "B7:B8", color_string)
    set_background_color(ws, "C15:C16", color_string)
    set_background_color(ws, "C18:C19", color_string)

    color_string = 'DBEEF4'  # light blue color hex string
    set_background_color(ws, "F14:F18", color_string)
    set_background_color(ws, "G14:I14", color_string)
    set_background_color(ws, "G18:I18", color_string)
    set_background_color(ws, "J14:J18", color_string)

    color_string = 'C3D69B'  # light green color hex string
    set_background_color(ws, "G15:I17", color_string)
    color_string = 'D99694'  # light red color hex string
    set_background_color(ws, "H16:H16", color_string)
    set_background_color(ws, "C17:C17", color_string)


def ctf_sheet_initiate(ws):
    "Initiate CTF sheet"
    for row in CTF_INITIATE_TABLE:
        ws.append(row)
    for idx, roi in enumerate(ROI_table):
        if idx <= 8:
            ws["B{}".format(23+idx)] = roi
        else:
            ws["E{}".format(23+idx-9)] = roi
    for idx, tgt in enumerate(TGT_FREQ):
        if idx <= 8:
            ws["C{}".format(23+idx)] = tgt
        else:
            ws["F{}".format(23+idx-9)] = tgt
    # yello FFFF00
    set_background_color(ws, "A2:Z2", 'FFFF00')
    set_background_color(ws, "A5:Z5", 'FFFF00')
    set_background_color(ws, "A9:Z12", 'FFFF00')
    set_background_color(ws, "A18:Z19", 'FFFF00')
    set_background_color(ws, "B23:B23", 'FFFF00')
    set_background_color(ws, "B26:B26", 'FFFF00')
    set_background_color(ws, "B30:B31", 'FFFF00')
    set_background_color(ws, "E23:E24", 'FFFF00')
    set_background_color(ws, "E30:E31", 'FFFF00')
    # blue B7DEE8
    set_background_color(ws, "C23:C31", 'B7DEE8')
    set_background_color(ws, "F23:F31", 'B7DEE8')
    # red D99694
    set_background_color(ws, "D23:D31", 'D99694')
    set_background_color(ws, "G23:G31", 'D99694')
    ws['C21'] = "Tgt Freq."
    ws['F21'] = "Tgt Freq."
    ws['B22'] = "ROI"
    ws['E22'] = "ROI"
    ws['D22'] = "MTF"
    ws['G22'] = "MTF"
    ws['F22'] = "cy/mm"
    ws['C22'] = "cy/mm"
    set_allborder(ws, "B21:G31")


def ctf_dealwith(ws):
    for idx in range(EY_FOLDERS_NUM):
        ws.cell(column=3+idx, row=1, value=EY_FOLDERS[idx])
        filefullpath = EY_FOLDERS[idx] + "/mtf/" + EY_FOLDERS[idx] \
            + "-H-MTFOUT.txt"
        with open(filefullpath, "r") as f:
            for line in f:
                roi, mtf = line.split('=')
                row_idx = ROI_table.index(roi) + 2
                ws.cell(column=3+idx, row=row_idx, value=float(mtf))
        filefullpath = EY_FOLDERS[idx] + "/mtf/" + EY_FOLDERS[idx] \
            + "-V-MTFOUT.txt"
        with open(filefullpath, "r") as f:
            for line in f:
                roi, mtf = line.split('=')
                row_idx = ROI_table.index(roi) + 2
                ws.cell(column=3+idx, row=row_idx, value=float(mtf))


def ctf_linechart(ws):
    chart1 = LineChart()
    chart1.title = "Line Chart"
    chart1.style = 2
    chart1.y_axis.title = ''
    chart1.x_axis.title = ''

    for idx in range(EY_FOLDERS_NUM):
        # Select all data include title
        data = Reference(ws, min_col=3+idx, min_row=1, max_row=19,
                         max_col=3+idx)
        chart1.add_data(data, titles_from_data=True)
        s1 = chart1.series[idx]
        s1.marker.symbol = "triangle"
        s1.marker.graphicalProperties.solidFill = "FF0000"  # Marker filling
        # Marker outline
        s1.marker.graphicalProperties.line.solidFill = "FF0000"
        s1.graphicalProperties.line.noFill = False
    ws.add_chart(chart1, "H5")


if __name__ == '__main__':
    wb = excel_create()
    create_working_sheets(wb)
    active_ws = wb[SHEET_SFR]
    sfr_sheet_initiate(active_ws)

    active_ws = wb[SHEET_SFR_2]
    sfr2_sheet_initiate(active_ws)

    active_ws = wb[SHEET_SFR]
    sfr_handle(active_ws)

    copy_sfr_to_sfr2(wb)
    active_ws = wb[SHEET_SFR_2]
    sfr2_linechart(active_ws)

    active_ws = wb[SHEET_CTF]
    ctf_sheet_initiate(active_ws)
    ctf_dealwith(active_ws)
    ctf_linechart(active_ws)
    excel_save(wb, FILENAME)
